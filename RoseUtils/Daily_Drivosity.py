import pandas as pd
import sqlite3
import openpyxl
from xlsxwriter.workbook import Workbook
import traceback
from os import remove
from time import sleep
def run(self, downloadsfolder, database, output):
    try:
        ######### USER CONFIG SECTION ############
        DOWNLOADS_FOLDER = downloadsfolder
        DATABASE_FILE = database
        OUTPUT_FILE = database[:database.rfind("/")]+"/Drivosity.xlsx"
        ######### END USER CONFIG   ##############

        self.ui.outputbox.setText("Running Daily Drivosity")


        ######Setup for pandas, SQLite and openpyxl
        conn = sqlite3.connect(DATABASE_FILE)
        cursor = conn.cursor()
        df = pd.read_excel(DOWNLOADS_FOLDER+"report.xlsx",usecols="D,F,G,U",skiprows = 1,header=0)
        wb = openpyxl.load_workbook(DOWNLOADS_FOLDER+"report.xlsx")
        ws = wb.active


        ##### Create pandas DataFrame
        df["Name"] = df["First Name"] +" " + df["Last Name"]
        df = df.drop(["First Name","Last Name"], axis=1 )
        df = df.rename(columns={"Store #":"store","DriveScore":"score","Name":"name"})


        ####Read the date from drivosity report and create list for dataframe, 
                                                        #then add to dataframe
        date = ws['a1'].value #type:ignore
        date = date[date.find("-")+1:-1]
        date = date.replace("/",".")
        date = [date] * len(df)
        df["date"] = date
        df = df[['date','store','name','score']]
        df.to_sql('drivosity',conn,if_exists = 'append',index=False)
        conn.commit()

        #####Export database to excel file
        workbook = Workbook(OUTPUT_FILE)
        worksheet = workbook.add_worksheet()
        cursor.execute("select * from drivosity")
        mysel=cursor.execute("select * from drivosity")
        for i, row in enumerate(mysel):
            for j, value in enumerate(row):
                worksheet.write(i, j, row[j])
        workbook.close()

        conn.close()
        if self.check_delete:
            remove(DOWNLOADS_FOLDER+"report.xlsx")

    except:
        self.ui.outputbox.append("ENCOUNTERED ERROR")
        self.ui.outputbox.append("Please send the contents of this box to Justin")
        self.ui.outputbox.append(traceback.format_exc())

