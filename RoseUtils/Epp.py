# type: ignore
from striprtf.striprtf import rtf_to_text
from os import listdir, remove
from openpyxl import Workbook
from openpyxl.styles import Alignment
import traceback
from time import sleep

def run(self):
    try:
        zeroes = ['4045','4052','4104','4113','4120','4121','4250','3064']
        self.ui.outputbox.setText("Running EPP")
        sleep(.01)
        filelist = [] #create list of files to loop through -> filelist
        for file in listdir(self.zocdownloadfolder):
            filelist.append(file)
        filelist = [self.zocdownloadfolder + file for file in filelist]

        if self.rcp:
            STORECOL = {"01740":2,"01743":4,"02172":6,"02174":8,"02236":10,"02272":12,"02457":14,"02549":16,"02603":18,"02953":20,"03498":22,"04778":24}
        else:
            STORECOL = {"2208":2,"2306":4,"2325":6,"2478":8,"2612":10,"2618":12,"2687":14,"2921":16,"3015":18,"3130":20,"3479":22,"4405":24}
                    #  
        PRODUCTS = {"10\" Dough":"1075",
                    "12\" Dough":"1076",
                    "14\" Dough":"1080",
                    "16\" Dough":"1082",
                    "Wings":"1092",
                    "Poppers":"1093",
                    "Chicken":"1095",}
        
        wb = Workbook()
        ws = wb.create_sheet('Sheet', 0) # type: ignore
        def dictindex(search_key, dict):
            return list(dict.values()).index(search_key) + 3


        def openrtf(filee): #Call this to open an rtf file with the filepath in the thing.
                        #Will return a list of strings for each line of the file
            with open(filee, 'r') as filee:
                text = filee.read()
            text = rtf_to_text(text)
            textlist = text.splitlines()
            return textlist #returns a list containing entire RTF file

        targets = [file for file in filelist if "INVTAR" in file]
        values = [file for file in filelist if "INVVAL" in file]
        
        for file in targets:
            file = openrtf(file)
            store = file[1][83:88].strip()
            columnnum = STORECOL[store] +1
            for line in file:
                if line[4:8] in PRODUCTS.values():
                    rownum = dictindex(line[4:8],PRODUCTS)
                    ws.cell(row = rownum, column = columnnum, value = float(line[116:125].strip())) 
                
            


        for file in values:
            file = openrtf(file)
            store = file[1][83:88].strip()
            columnnum = STORECOL[store]
            for line in file:
                try: 
                    if line[2:6] in zeroes:
                        amount = float(line[48:54].strip())
                        if amount > 0:
                            itemname = line[8:34].strip()
                            self.ui.outputbox.append(f"Store: {store}, Item: {itemname}, {amount}")
                except:
                    pass
                if line[2:6] in PRODUCTS.values():
                    rownum = dictindex(line[2:6],PRODUCTS)
                    ws.cell(row = rownum, column = columnnum, value = float(line[48:54].strip()))



        #set up headers
        for item in STORECOL.items():
            ws.cell(row = 1, column = item[1], value = item[0])
            ws.cell(row = 2, column = item[1], value = "On Hand")
            ws.cell(row = 2, column = item[1]+1, value = "Yield")
        for i, item in enumerate(PRODUCTS.keys()):
            ws.cell(row = i+3,column = 1, value = item)
        ws.merge_cells("b1:c1")
        ws.merge_cells("d1:e1")
        ws.merge_cells("f1:g1")
        ws.merge_cells("h1:i1")
        ws.merge_cells("j1:k1")
        ws.merge_cells("l1:m1")
        ws.merge_cells("n1:o1")
        ws.merge_cells("p1:q1")
        ws.merge_cells("r1:s1")
        ws.merge_cells("t1:u1")
        ws.merge_cells("v1:w1")
        ws.merge_cells("x1:y1")
        rows = range(1, 2)
        columns = range(1, 50)
        for row in rows:
            for col in columns:
                ws.cell(row, col).alignment = Alignment(horizontal='center')




        wb.save(self.outputfolder+"EPP.xlsx")
        self.ui.outputbox.append("Done!")
        sleep(.01)
        if self.check_delete:
            for file in targets:
                remove(file)
            for file in values:
                remove(file)

    except:
        self.ui.outputbox.append("ENCOUNTERED ERROR")
        self.ui.outputbox.append("Please send the contents of this box to Justin")
        self.ui.outputbox.append(traceback.format_exc())






