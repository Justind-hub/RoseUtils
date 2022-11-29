from os import listdir, remove, startfile
from openpyxl import Workbook
from RoseUtils.rtf_reader import read_rtf
import traceback


def run(self):    
    try:
        def decimal_time(time:str)->float: 
            '''Pass in a string "2:38pm" and return a float value of 14.633_'''
            h = int(time[0:2])
            if h == 12 and time[-2:] == "am": h = 0
            if time[-2:].lower() == "pm":
                h +=12
            m = int(time[3:5])/60
            if h == 24: h = 12
            return h+m

        PATH = self.zocdownloadfolder
        wb = Workbook()
        ws = wb.active
        r = 1

        for file in listdir(PATH):
            if "ORDDTL" not in file: continue
            report = read_rtf(PATH+file)
            if self.check_delete: remove(PATH+file)
            store = int(report[0][82:89].strip())
            print(len(report[8]))
            report = [line for line in report if len(line) == 150 or len(line) == 124 or len(line) == 30 or len(line) == 152 and "ExtOrdID" not in line]
            date = ""
            self.append_text.emit(f"Currently running store {store}...")
            for line in report:
                if len(line) == 30: date = line.strip()
                if len(line)>5: 
                    if line[5] == "C" and line[6] != "a":
                        if decimal_time(line[84:96].strip()) > 21.66: 
                            ws.cell(row = r, column = 1, value = store)
                            ws.cell(row = r, column = 2, value = date)
                            ws.cell(row = r, column = 3, value = line[87:94])
                            ws.cell(row = r, column = 4, value = line[2:6]) 
                            ws.cell(row = r, column = 5, value = float(line[140:146].strip())) 
                            r+=1                
        wb.save(self.outputfolder+"eveningsales.xlsx")
        startfile(self.outputfolder+"eveningsales.xlsx")
        self.append_text.emit(f"Report complete, opening file now")
        
    except:
        self.append_text.emit("ENCOUNTERED ERROR")
        self.append_text.emit("Please send the contents of this box to Justin")
        self.append_text.emit(traceback.format_exc())
        pass