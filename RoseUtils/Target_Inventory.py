#from striprtf.striprtf import rtf_to_text
# type: ignore
from RoseUtils.rtf_reader import read_rtf as openrtf
import os
from os.path import exists
from PyQt5.QtWidgets import QApplication
import pandas as pd
from time import sleep
from openpyxl import Workbook 
from openpyxl.styles import Alignment, Font
from datetime import datetime
from time import perf_counter
import traceback

def run(self):
    try:
        
        start = perf_counter()
        self.append_text.emit("Running Target Inventory")
        wb= Workbook()
        ws = wb.create_sheet('Sheet', 0)  # type: ignore
        if self.rcp:
            STORECOL = {"1740":1,"1743":2,"2172":3,"2174":4,"2236":5,"2272":6,"2457":7,"2549":8,"2603":9,"2953":10,"3498":11,"4778":12}
        else:
            STORECOL = {"2208":1,"2306":2,"2325":3,"2478":4,"2612":5,"2618":6,"2687":7,"2921":8,"3015":9,"3130":10,"3479":11,"4405":12}
        VARIANCEAMOUNT = 10
        storelist = ["1740","1743","2172","2174","2236","2272","2457","2549","2603","2953","3498","4778"] if self.rcp else \
                    ["2208","2306","2325","2478","2612","2618","2921","3015","3130","3479","4405"]

        #openrtf file
        '''def openrtf(file): #Call this to open an rtf file with the filepath in the thing.
                            #Will return a list of strings for each line of the file
            with open(file, 'r') as file:
                text = file.read()
            text = rtf_to_text(text)
            textlist = text.splitlines()
            return textlist #returns a list containing entire RTF file'''

        def left(s, amount): #litterally just the "left" function from excel
            return s[:amount]

        filelist = [] #create list of files to loop through -> filelist
        for file in os.listdir(self.zocdownloadfolder):
            if file.startswith("INVTAR"):
                filelist.append(file)
        f = 0
        for each in filelist:
                
            #clean the file
            textlist = openrtf(self.zocdownloadfolder+each)
            headerstart = left(textlist[1],10)
            i = 10
            while i < len(textlist):
                if left(textlist[i],10) == headerstart:
                    del textlist[i-2:i+3]
                    i += 4
                else:
                    i += 1
            while("" in textlist) :
                textlist.remove("")
            del textlist[-22:]

            i = 3
            while i < len(textlist):
                try:
                    x = int(left(textlist[i], 7))
                except:
                    del textlist[i]
                    i-=1
                i +=1

            #Date and store number
            store = textlist[0]
            store = store[82:90].strip()
            store = store[-4:]
            try:
                storelist.remove(store.strip())
            except:
                pass
            self.append_text.emit(f"Target Inventory: Running store {store}")
            #store = int(store)
            date = textlist[2]
            date = date.strip()
            del textlist[0:4]
            col = STORECOL[store]
            #create 3 lists of item#, item name and variance

            item = []
            dvariance = []
            
            for i, line in enumerate(textlist):
                item.append(line[11:34].strip())
                dvariance.append(float(line[105:115].strip()))


            df = pd.DataFrame(data={'Item':item,'$ Variance':dvariance})
            #df.sort_values('$ Variance', inplace=True)
            df.drop(df[(df['$ Variance'] > -VARIANCEAMOUNT) & (df['$ Variance'] < VARIANCEAMOUNT)].index, inplace = True)
            df['Variance'] = df['Item'] + " " +df['$ Variance'].map(str)
            df.sort_values(by='$ Variance', key=abs, inplace=True, ascending=False)
            df.reset_index(inplace=True)
            df.drop(['Item','$ Variance','index'], axis=1, inplace=True)
            ws.cell(row = 2, column = col, value = int(store))
            
            #Write variances to cells
            i = 0
            variances = df['Variance'].to_list()
            while i < len(variances):
                ws.cell(column = col, row = i+3, value = variances[i])
                i+=1
            

            f +=1
            if self.check_delete:
                os.remove(self.zocdownloadfolder+each)




        for col in ws.columns:
            column = col[0].column_letter # Get the column name
            maxl = 8
            for cell in col:
                x = str(cell.internal_value)
                if len(x) > maxl:
                    maxl = len(x)
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[column].width = maxl/1.09
            if maxl == 8:
                ws.cell(row = 3, column = col[0].column, value = "No Variances over $"+str(VARIANCEAMOUNT)).font = Font(bold=True)
                ws.column_dimensions[column].width = 20

        ws.cell(row = 1, column = 1, value = date).alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:L1")
        end_time = perf_counter()
        if self.rcp:
            wb.save(self.outputfolder+"Inventory Target "+datetime.now().strftime("%m.%d.%y")+".xlsx")
        else:
            wb.save(self.outputfolder+"CCDInventory Target "+datetime.now().strftime("%m.%d.%y")+".xlsx")
        self.append_text.emit("Done!")
        self.append_text.emit(f"Target Inventory: Completed in {round(end_time - start,2)} seconds")    
        self.append_text.emit(f"Missing stores {storelist}")
    except:
        self.append_text.emit("ENCOUNTERED ERROR")
        self.append_text.emit("Please send the contents of this box to Justin")
        self.append_text.emit(traceback.format_exc())


if __name__ == '__main__':
    run()