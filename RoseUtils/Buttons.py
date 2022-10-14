from RoseUtils.Main_UI import Ui_RoseUtils

from PyQt5 import QtWidgets as qtw
from PyQt5 import QtCore as qtc
from PyQt5 import QtGui
from PyQt5.QtWidgets import  QMessageBox, QFileDialog # Change to * if you get an error
from RoseUtils import Daily_DOR_Breaks, New_Hire, Target_Inventory
from RoseUtils import Weekly_DOR_CSC, Weeklycompfull, Daily_Drivosity
from RoseUtils import Epp, Comments, Release, gm_Target_inv, gm_weeklycomp
from RoseUtils import export_SQL,gm_on_hands, updater, costreport
import winsound
import time

import os
from os.path import exists
from subprocess import Popen, PIPE
import logging
#import PyPDF2
#from PyPDF2 import PdfFileReader, PdfFileWriter


class Buttons():
    def pa_promo(_, self):
        import pandas as pd
        from openpyxl import Workbook
        import numpy as np

        
        costlist, check = QFileDialog.getOpenFileNames(None, "QFileDialog.getOpenFileName()",
                        self.downloadfolder,"Excel Files (*.xlsx)")
        if check:
            if len(costlist) == 1:
                PATH = costlist[0]
            else:
                self.popup("Select exactly 1 report1",
                              QMessageBox.Warning,"Error")
                return




        wb = Workbook()
        ws = wb.active
        if self.rcp:
            STORES = [1740,2172,2236,2272,2549,2953,4778,1743,2174,2457,2603,3498]
        else:
            STORES = [2208,2306,2325,2478,2612,2618,2687,2921,3015,3130,3479,4405]


        df = pd.read_excel(PATH, skiprows=11,usecols=[0,5,7])
        data = df.to_numpy()

        for i, store in enumerate(STORES):
            r = 3
            tms = 0
            complete = 0
            ws.cell(row = 1, column = i+1,value = store)
            for row in data:
                fname = row[0][row[0].find(" ")+1:]
                lname = row[0][:row[0].find(" ")]
                if str(store) in row[2]:
                    tms +=1
                    if row[1] == 100: 
                        complete += 1
                    else:
                        ws.cell(row = r, column = i+1, value = f"{fname} {lname}")
                        r += 1
            ws.cell(row = 2, column = i+1, value = str(round((complete/tms) * 100,1)) + "%")

        

        wb.save(self.outputfolder + "PA Promo Report.xlsx")
        if self.check_delete:
            os.remove(PATH)
        os.startfile(self.outputfolder + "PA Promo Report.xlsx")
        del(STORES,df,data,wb)

