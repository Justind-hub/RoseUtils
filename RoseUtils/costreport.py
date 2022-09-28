from striprtf.striprtf import rtf_to_text
from openpyxl import Workbook
def openrtf(file): #Call this to open an rtf file with the filepath in the thing.
                        #Will return a list of strings for each line of the file
        with open(file, 'r') as file:
            text = file.read()
        text = rtf_to_text(text)
        textlist = text.splitlines()
        return textlist #returns a list containing entire RTF file

def run(self):
    textlist = openrtf(self.cost_report_list[0])
    date1 = textlist[3].strip()
    textlist = [line for line in textlist if len(line) == 119]

    itemNums1 = [line[3:9].strip() for line in textlist]
    items1 = [line[9:43].strip() for line in textlist]
    used1 = [line[77:89].strip() for line in textlist]
    costOfUsed1 = [line[89:103].strip() for line in textlist]
    cost1 = []
    for a, b in zip(used1, costOfUsed1):
        if float(b) != 0:
            cost1.append(float(b)/float(a))
        else:
            cost1.append(0)

    textlist = openrtf(self.cost_report_list[1])
    date2 = textlist[3].strip()
    textlist = [line for line in textlist if len(line) == 119]

    itemNums2 = [line[3:9].strip() for line in textlist]
    items2 = [line[9:43].strip() for line in textlist]
    used2 = [line[77:89].strip() for line in textlist]
    costOfUsed2 = [line[89:103].strip() for line in textlist]
    cost2 = []
    for a, b in zip(used2, costOfUsed2):
        if float(b) != 0:
            cost2.append(float(b)/float(a))
        else:
            cost2.append(0)



    wb = Workbook()
    ws = wb.active

    ws.cell(row=1,column=1,value="Item Number")
    ws.cell(row=1,column=2,value="Item")
    ws.cell(row=1,column=3,value="Used")
    ws.cell(row=1,column=4,value="Cost/item 1")
    ws.cell(row=1,column=5,value="Cost/item 2")
    ws.cell(row=1,column=6,value="Cost 1")
    ws.cell(row=1,column=7,value="Cost 2")
    ws.cell(row=1,column=8,value="Difference")


    for i in range(len(items1)):
        ws.cell(row=i+2,column=1,value=itemNums1[i])
        ws.cell(row=i+2,column=2,value=items1[i])
        ws.cell(row=i+2,column=3,value=used1[i])
        ws.cell(row=i+2,column=4,value=round(cost1[i],2))
        
        try:
            costa2 = round(cost2[itemNums2.index(itemNums1[i])],2)
            ws.cell(row=i+2,column=5,value=costa2)
        except:
            pass
        ws.cell(row=i+2,column=6,value=round(float(used1[i]) * float(cost1[i]),2))
        ws.cell(row=i+2,column=7,value=round(float(used1[i]) * float(costa2),2))
        if ws.cell(row=i+2,column=6).value == 0 or ws.cell(row=i+2,column=7).value == 0:
            pass
        else:
            ws.cell(row=i+2,column=8,value =round((float(used1[i]) * float(costa2)) - (float(used1[i]) * float(cost1[i])),2))

    ws.cell(column=10,row=2,value="Date Range 1:")
    ws.cell(column=10,row=3,value="Date Range 2:")
    ws.cell(column=10,row=5,value="Total Difference:")
    ws.cell(column=11,row=2,value=date1)
    ws.cell(column=11,row=3,value=date2)
    ws.cell(column=11,row=5,value="=sum(h2:h500)")

    wb.save(self.outputfolder + "Commodity Price Change Report.xlsx")
    self.ui.outputbox.setText("Running Commodity Price Change Report")
    self.ui.outputbox.append("Report ran, outputted to your output folder")

