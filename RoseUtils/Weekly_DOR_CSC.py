# type: ignore
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from striprtf.striprtf import rtf_to_text 
import traceback
from time import sleep

def run(self, zocdownload, output):
    try:
            

    

        self.ui.outputbox.setText("Running Weekly DOR/CSC")
        sleep(.01)
        wb = Workbook()

        #get desktop folder to save to
        
        if self.rcp:
            storecol = {"1740":2,"1743":3,"02172":4,"2174":5,"02236":6,"2272":7,"02457":8,"2549":9,"02603":10,"2953":11,"3498":12,"04778":13}
        else:
            storecol = {"2208":2,"2306":3,"2325":4,"2478":5,"2612":6,"2618":7,"2687":8,"2921":9,"3015":10,"3130":11,"3479":12,"4405":13}


        def openrtf(file): #Call this to open an rtf file with the filepath in the thing.
                            #Will return a list of strings for each line of the file
            with open(file, 'r') as file:
                text = file.read()
            text = rtf_to_text(text)
            textlist = text.splitlines()
            return textlist #returns a list containing entire RTF file

        def daypart(part): #Returns first instance of the line containing day part sales data
            filter_object = filter(lambda a: part in a, textlist)
            return list(filter_object)[0]

        def findline(text, start): #finds line containing text
            i = start
            while i <len(textlist):
                if textlist[i].find(text) == -1:
                    i = i+1
                    continue
                else:
                    break
            return i

        def left(s, amount): #litterally just the "left" function from excel
            return s[:amount]



        filelist = [] #create list of files to loop through -> filelist
        for file in os.listdir(zocdownload):
            if file.startswith("ODMDOR"):
                filelist.append(file)

        ##########################################################################
        #begining of main loop
        f = 0
        for each in filelist:
            missedbreaks = 0
            textlist = openrtf(zocdownload + filelist[f])
            lunch = daypart("Lunch")
            afternoon = daypart("Afternoon")
            dinner = daypart("Dinner")
            evening = daypart("Evening")
            
            ##########################################################################
            #Get store number and date
            store = textlist[1]
            store = store[store.find("RESTAURANT")+11:store.find("RESTAURANT")+16].strip()
            col = storecol[store]
            date = textlist[4].strip()
            #create new worksheet with name {date}
            try:
                if wsname == date:
                    pass
            except:
                wsname = date.replace("/",".")
                ws = wb.create_sheet(wsname, 0)
                ws.title = wsname

            #########################################################################
            #remove header from each page
            headerstart = left(textlist[1],10)
            i = 10
            while i < len(textlist):
                if left(textlist[i],10) == headerstart:
                    del textlist[i-2:i+4]
                    i += 4
                else:
                    i += 1
            while("" in textlist) :
                textlist.remove("")


            
            ##########################################################################
            #Get total day CSC
            cscline = findline("Order Summary",30)+3
            cscline = textlist[cscline]
            csc = cscline[127:]
            csc = csc.strip()

            ##########################################################################
            # Find and print OTD and CSC for each daypart
            lunchcsc = str(int(float(lunch[-6:])))
            lunchotd = lunch[lunch.find(":")+1:lunch.find(":")+6]
            afternooncsc = str(int(float(afternoon[-6:])))
            afternoonotd = afternoon[afternoon.find(":")+1:afternoon.find(":")+6]
            dinnercsc = str(int(float(dinner[-6:])))
            dinnerotd = dinner[dinner.find(":")+1:dinner.find(":")+6]
            eveningcsc = str(int(float(evening[-6:])))
            eveningotd = evening[evening.find(":")+1:evening.find(":")+6]

            
            
            ws.cell(row= 2, column = 1, value = "Store #:")
            ws.cell(row= 3, column = 1, value = "CSC:")
            ws.cell(row= 4, column = 1, value = "Lunch:")
            ws.cell(row= 5, column = 1, value = "Afternoon:")
            ws.cell(row= 6, column = 1, value = "Dinner:")
            ws.cell(row= 7, column = 1, value = "Evening:")
            ws.cell(row= 8, column = 1, value = "Voids")
            ws.cell(row= 8, column = col, value = float(textlist[findline("Void",5)][30:38].strip()))
            ws.cell(row = 2, column = col, value = int(store))
            ws.cell(row = 3, column = col, value = csc+"%")    
            ws.cell(row = 4, column = col, value = lunchotd+" "+lunchcsc+"%")
            ws.cell(row = 5, column = col, value = afternoonotd+" "+afternooncsc+"%")
            ws.cell(row = 6, column = col, value = dinnerotd+" "+dinnercsc+"%")
            ws.cell(row = 7, column = col, value = eveningotd+" "+eveningcsc+"%")
            
            if self.check_delete:
                os.remove(zocdownload + each)
            f += 1

        for col in ws.columns:
            column = col[0].column_letter # Get the column name
            ws.column_dimensions[column].width = 13
            for cell in col:
                cell.alignment = Alignment(horizontal="center")

        ws.cell(row=1, column = 1, value = date)
        ws.merge_cells("A1:"+column+"1")
        
        wb.save(output+date.replace("/",".")+".xlsx")

    except:
        self.ui.outputbox.append("ENCOUNTERED ERROR")
        self.ui.outputbox.append("Please send the contents of this box to Justin")
        self.ui.outputbox.append(traceback.format_exc())


if __name__ == '__main__':
    run()