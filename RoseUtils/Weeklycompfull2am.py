# type: ignore
#from striprtf.striprtf import rtf_to_text
from RoseUtils.rtf_reader import read_rtf as openrtf
import os
from os.path import exists

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from datetime import datetime
import traceback
from time import sleep, perf_counter


def run(self):
    try:
        start = perf_counter()
        zocdownload = self.zocdownloadfolder
        output = self.outputfolder
        self.append_text.emit("Running Weekly Comp")
        sleep(.01)


        leftborder = Border(left=Side(style='thick'))
        rightborder = Border(right=Side(style='thick'))
        bottomborder = Border(bottom=Side(style='thick'))
        allborder = Border(bottom=Side(style='thick'),right=Side(style='thick'),left=Side(style='thick'))
        fill3 = PatternFill(start_color='00EEECE1', end_color='00EEECE1', fill_type='solid')
        fill12 = PatternFill(start_color='00E4DFEC', end_color='00E4DFEC', fill_type='solid')
        fill7 = PatternFill(start_color='00FDE9D9', end_color='00FDE9D9', fill_type='solid')



        wb= Workbook()
        ws = wb.create_sheet('Weekly Comp', index = 0)
        if self.rcp:
            STORECOL = {"1740":24,"1743":32,"2172":28,"2174":0,"2236":4,"2272":36,"2457":40,"2549":8,"2603":44,"3498":16,"4778":20,"2953":12}
            stores = [2174,2236,2549,2953,3498,4778,1740,2172,1743,2272,2457,2603]
        else:
            STORECOL = {"2208":0,"2306":4,"2325":8,"2478":12,"2612":16,"2618":20,"2687":24,"2921":28,"3015":32,"3130":36,"3479":40,"4405":44,"5293":48,"5424":52}
            stores = [2208,2306,2325,2478,2612,2618,2687,2921,3015,3479,4405,5293,5424]
        storelist = ["1740","1743","2172","2174","2236","2272","2457","2549","2603","2953","3498","4778"] if self.rcp else \
                    ["2208","2306","2325","2478","2612","2618","2687","2921","3015","3130","3479","4405","5293","5296"]
        weekday = [0,19,38,57,76,95,114]

        #set headers at each store
        
        worksheets = [ws]


        

        #openrtf file
        '''def openrtf(file): #Call this to open an rtf file with the filepath in the thing.
                            #Will return a list of strings for each line of the file
            with open(file, 'r') as file:
                text = file.read()
            text = rtf_to_text(text)
            textlist = text.splitlines()
            return textlist #returns a list containing entire RTF file'''

        def left(s, amount): #litterally just the "left" function from excel
            return s[:amount]

        filelist = [] #create list of files to loop through -> filelist
        for file in os.listdir(zocdownload):
            if file.startswith("DSHWKC"):
                filelist.append(file)
                
        def getday(text, a): #send text file and starting line for the day of the week, returns dataframe
            drivers = []
            starthour = []
            instores = []
            deliveries = []
            products = []
            for i,line in enumerate(text):
                drivers.append(round(float(line[a+8:a+13].strip()),1))
                instores.append(round(float(line[a+15:a+21].strip()),1))
                products.append(int(line[a+22:a+25].strip()))
                deliveries.append(int(line[a:a+2].strip()))
                starthour.append(int(line[:2]))
            return(pd.DataFrame({"Time":starthour,'Deliveries':deliveries,'Drivers':drivers,'Products':products,'Instores':instores}))

        #returns list of short hours of day df passed into function
        def getshort(df):
            df['delpdr'] = df['Deliveries'] / df['Drivers']
            df['prodpin'] = df['Products'] / df['Instores']
            export = []
            end = []
            for index, row in df.iterrows():
                if int(row[0]) == 12:
                    end.append("1")
                else:
                    end.append(str(int(row[0] + 1)))
                if row['delpdr'] > 3:
                    if row['prodpin'] > 15:
                        export.append("b") #write both dr and in to export
                    else:
                        export.append("d") #write just dr to export
                elif row['prodpin'] > 15:
                    export.append("i") #write just in to export
                else: 
                    export.append("")
            df['export'] = export
            df['endtime'] = end
            df['Time'] = df['Time'].map(str)+"-"+df['endtime']

            #delete these 4 rows to export entire day
            #df['export'].replace('', np.nan, inplace=True)
            #df.dropna(subset=['export'], inplace=True)
            #df.reset_index(inplace=True)
            #df.drop(['Time','Deliveries','Drivers','Products','Instores','delpdr','prodpin','index'], axis=1, inplace=True)
            df.drop(['delpdr','prodpin','endtime'], axis=1, inplace=True)
            df = df[['Time','Deliveries','Drivers','Instores','Products','export']]
            list1 = df.values.tolist()    
            return(df)
            ############# color = EEECE1
        def printall(wday, df):
            rows123 = df.shape[0]
            
            i = 0
            #col[0].column_letter is the column letter
            ws.cell(row = i+2+wday, column = col+2, value = 'Del').border = bottomborder
            ws.cell(row = i+2+wday, column = col+3, value = 'DR').border = bottomborder
            ws.cell(row = i+2+wday, column = col+4, value = 'IN').border = bottomborder
            ws.cell(row = i+2+wday, column = col+5, value = 'PROD').border = bottomborder
            while i < rows123:
                if df.iloc[i]['export'] == "b":
                    ws.cell(row = i+3+wday, column = col+2, value = df.iloc[i]['Deliveries']).font = Font(bold=True, color="00FF0000")
                    ws.cell(row = i+3+wday, column = col+2).border = leftborder
                    ws.cell(row = i+3+wday, column = col+3, value = df.iloc[i]['Drivers']).font = Font(bold=True, color="00FF0000")
                    ws.cell(row = i+3+wday, column = col+4, value = df.iloc[i]['Instores']).font = Font(bold=True, color="00FF0000")
                    ws.cell(row = i+3+wday, column = col+5, value = df.iloc[i]['Products']).font = Font(bold=True, color="00FF0000")
                    ws.cell(row = i+3+wday, column = col+5).border = rightborder
                elif df.iloc[i]['export'] == "d":
                    ws.cell(row = i+3+wday, column = col+2, value = df.iloc[i]['Deliveries']).font = Font(bold=True, color="00FF0000")
                    ws.cell(row = i+3+wday, column = col+2).border = leftborder
                    ws.cell(row = i+3+wday, column = col+3, value = df.iloc[i]['Drivers']).font = Font(bold=True, color="00FF0000")
                    ws.cell(row = i+3+wday, column = col+4, value = df.iloc[i]['Instores'])
                    ws.cell(row = i+3+wday, column = col+5, value = df.iloc[i]['Products'])
                    ws.cell(row = i+3+wday, column = col+5).border = rightborder
                elif df.iloc[i]['export'] == "i":
                    ws.cell(row = i+3+wday, column = col+2, value = df.iloc[i]['Deliveries'])
                    ws.cell(row = i+3+wday, column = col+2).border = leftborder
                    ws.cell(row = i+3+wday, column = col+3, value = df.iloc[i]['Drivers'])
                    ws.cell(row = i+3+wday, column = col+4, value = df.iloc[i]['Instores']).font = Font(bold=True, color="00FF0000")
                    ws.cell(row = i+3+wday, column = col+5, value = df.iloc[i]['Products']).font = Font(bold=True, color="00FF0000")
                    ws.cell(row = i+3+wday, column = col+5).border = rightborder
                else:
                    ws.cell(row = i+3+wday, column = col+2, value = df.iloc[i]['Deliveries'])
                    ws.cell(row = i+3+wday, column = col+2).border = leftborder
                    ws.cell(row = i+3+wday, column = col+3, value = df.iloc[i]['Drivers'])
                    ws.cell(row = i+3+wday, column = col+4, value = df.iloc[i]['Instores'])
                    ws.cell(row = i+3+wday, column = col+5, value = df.iloc[i]['Products'])
                    ws.cell(row = i+3+wday, column = col+5).border = rightborder
                i+=1

            



            


        for file in filelist:

            textlist = openrtf(zocdownload+file)
            textlist.insert(0,"")
            textlist.insert(5,"")
            if self.check_delete:
                os.remove(zocdownload+file)
            #set store number and date range, and save  
            store = textlist[1]
            store = store[82:90].strip()
            store = store[-4:]
            try:
                storelist.remove(store.strip())
            except:
                pass
            date = textlist[3]
            date = date.strip()
            col = STORECOL[store]
            #print(f"Loading store {store} from file {file}....")
            self.append_text.emit(f"Weekly Comp: Loading store {store}")
            sleep(.01)
            #clean up the file
            del textlist[50:]
            x = []
            for i, line in enumerate(textlist):
                if line[:5] == "     ":
                    x.append(i)
            x.reverse()

                #Delete useless lines, leaving data starting in row 5
            for y in x:
                del textlist[y]
            del textlist[:4]

            #create text1 and text2 for weekdays and weekend
            for i, x in enumerate(textlist):
                if len(x) == 0:
                    stoppoint = i
                    break

            text1 = textlist[:stoppoint]
            del textlist[:stoppoint+1]
            for i, x in enumerate(textlist):
                if len(x) == 0:
                    stoppoint = i
                    break
            text2 = textlist[:stoppoint]
            del textlist[:]
            #get each day as a df

            monday = getday(text1, 22)
            tuesday = getday(text1, 22+29)
            wednesday = getday(text1, 22+29+29)
            thursday = getday(text1, 22+29+29+29)
            friday = getday(text2, 22)
            saturday = getday(text2, 22+29)
            sunday = getday(text2, 22+29+29)
            days = [monday, tuesday, wednesday, thursday, friday, saturday, sunday]
            worksheets = [ws]
            for i,df in enumerate(days):
                df = getshort(df)

            for i,day in enumerate(days):
                printall(weekday[i],day)




        rows = range(1, 150)
        columns = range(1, 60)
        for row in rows:
            for col in columns:
                ws.cell(row, col).alignment = Alignment(horizontal='center')



        for col in ws.columns:
            column = col[0].column_letter # Get the column name
            ws.column_dimensions[column].width = 5
        

        for i in weekday:
            ws.cell(row = 1+i, column = 2, value = stores[1-1]).border = allborder
            ws.cell(row = 1+i, column = 6, value = stores[2-1]).border = allborder
            ws.cell(row = 1+i, column = 10, value = stores[3-1]).border = allborder
            ws.cell(row = 1+i, column = 14, value = stores[4-1]).border = allborder
            ws.cell(row = 1+i, column = 18, value = stores[5-1]).border = allborder
            ws.cell(row = 1+i, column = 22, value = stores[6-1]).border = allborder
            ws.cell(row = 1+i, column = 26, value = stores[7-1]).border = allborder
            ws.cell(row = 1+i, column = 30, value = stores[8-1]).border = allborder
            ws.cell(row = 1+i, column = 34, value = stores[9-1]).border = allborder
            ws.cell(row = 1+i, column = 38, value = stores[10-1]).border = allborder
            ws.cell(row = 1+i, column = 42, value = stores[11-1]).border = allborder
            ws.cell(row = 1+i, column = 42).alignment = Alignment(horizontal="center")
            ws.cell(row = 1+i, column = 46, value = stores[12-1]).border = allborder
            ws.cell(row = 1+i, column = 46).alignment = Alignment(horizontal="center")
            if not self.rcp:
                ws.cell(row = 1+i, column = 50, value = stores[13-1]).border = allborder
                ws.cell(row = 1+i, column = 50).alignment = Alignment(horizontal="center")
                ws.cell(row = 1+i, column = 54, value = stores[13-1]).border = allborder
                ws.cell(row = 1+i, column = 54).alignment = Alignment(horizontal="center")

            ws.cell(row=3+i, column = 1, value ="10-11")
            ws.cell(row=4+i, column = 1, value ="11-12")
            ws.cell(row=5+i, column = 1, value ="12-1")
            ws.cell(row=6+i, column = 1, value ="1-2")
            ws.cell(row=7+i, column = 1, value ="2-3")
            ws.cell(row=8+i, column = 1, value ="3-4")
            ws.cell(row=9+i, column = 1, value ="4-5")
            ws.cell(row=10+i, column = 1, value ="5-6")
            ws.cell(row=11+i, column = 1, value ="6-7")
            ws.cell(row=12+i, column = 1, value ="7-8")
            ws.cell(row=13+i, column = 1, value ="8-9")
            ws.cell(row=14+i, column = 1, value ="9-10")
            ws.cell(row=15+i, column = 1, value ="10-11")
            ws.cell(row=16+i, column = 1, value ="11-12")
            ws.cell(row=17+i, column = 1, value ="12-1")
            ws.cell(row=18+i, column = 1, value ="1-2")
            ii = 1
            lastcolnum = 58
            if self.rcp: lastcolnum = 50
            while ii < lastcolnum:
                ws.cell(row=8+i, column = ii).fill = fill3
                ws.cell(row=5+i, column = ii).fill = fill12
                ws.cell(row=12+i, column = ii).fill = fill7
                ii += 1


        templist = ["2","21","40","59","78","97","116"]

        ws.insert_rows(1)
        for i in templist:
            ws.merge_cells(f"B{i}:E{i}")
            ws.merge_cells(f"F{i}:I{i}")
            ws.merge_cells(f"J{i}:M{i}")
            ws.merge_cells(f"N{i}:Q{i}")
            ws.merge_cells(f"R{i}:U{i}")
            ws.merge_cells(f"V{i}:Y{i}")
            ws.merge_cells(f"Z{i}:AC{i}")
            ws.merge_cells(f"AD{i}:AG{i}")
            ws.merge_cells(f"AH{i}:AK{i}")
            ws.merge_cells(f"AL{i}:AO{i}")
            ws.merge_cells(f"AP{i}:AS{i}")
            ws.merge_cells(f"AT{i}:AW{i}")
            if not self.rcp:
                ws.merge_cells(f"AX{i}:BA{i}")
                ws.merge_cells(f"BB{i}:BE{i}")
        
        lastcol = "be"
        if self.rcp: lastcol = "aw"

        ws.merge_cells(f"a1:{lastcol}1")
        ws.merge_cells(f"a20:{lastcol}20")
        ws.merge_cells(f"a39:{lastcol}39")
        ws.merge_cells(f"a58:{lastcol}58")
        ws.merge_cells(f"a77:{lastcol}77")
        ws.merge_cells(f"a96:{lastcol}96")
        ws.merge_cells(f"a115:{lastcol}115")


        ws.cell(row=1, column = 1).font = Font(bold=True, size = 30)
        ws.cell(row=1, column = 1,value = "Monday").alignment = Alignment(horizontal="center")
        ws.cell(row=20, column = 1,value = "Tuesday").font = Font(bold=True, size = 30)
        ws.cell(row=39, column = 1,value = "Wednesday").font = Font(bold=True, size = 30)
        ws.cell(row=58, column = 1,value = "Thursday").font = Font(bold=True, size = 30)
        ws.cell(row=77, column = 1,value = "Friday").font = Font(bold=True, size = 30)
        ws.cell(row=96, column = 1,value = "Saturday").font = Font(bold=True, size = 30)
        ws.cell(row=115, column = 1,value = "Sunday").font = Font(bold=True, size = 30)

        end_time = perf_counter()
        if self.rcp:
            wb.save(output+"Weekly Comp "+datetime.now().strftime("%m.%d.%y")+".xlsx")
        else:
            wb.save(output+"CCDWeekly Comp "+datetime.now().strftime("%m.%d.%y")+".xlsx")
        self.append_text.emit(f"Weekly Comp: Completed in {round(end_time - start,2)} seconds")    
        self.append_text.emit(f"Missing stores {storelist}")
    except:
        self.append_text.emit("ENCOUNTERED ERROR")
        self.append_text.emit("Please send the contents of this box to Justin")
        self.append_text.emit(traceback.format_exc())

if __name__ == '__main__':
    run()