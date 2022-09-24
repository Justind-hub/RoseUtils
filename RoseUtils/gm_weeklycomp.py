from logging import exception
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, Color
import os
import traceback
from striprtf.striprtf import rtf_to_text
import math

def openrtf(file: str) -> list[str]: #Call this to open an rtf file with the filepath in the thing.
                    #Will return a list of strings for each line of the file
    with open(file, 'r') as file:
        text = file.read()
    text = rtf_to_text(text)
    textlist = text.splitlines()
    return textlist #returns a list containing entire RTF file


def findline(filel: list[str], string: str) -> int:
    for i, line in enumerate(filel):
        if string in line:
            return i
    raise Exception(f"'{string}' was not found in file provided!")
    return 0
    


def run(self, filelist):
    try:
        self.ui.outputbox.setText("Running Schedule History....")
        #Variables
        wb = Workbook()
        ws = wb.active
        HOURS = ['10-11','11-12','12-1','1-2','2-3','3-4','4-5','5-6','6-7','7-8','8-9','9-10']
        HOURSCOLS = [1,10,19,28,37,1,10,19,28,37,1,10,19,28,1,10,19,28]
        HOURSROWS = [5,5,5,5,5,19,19,19,19,19,35,35,35,35,49,49,49,49]
        DFORMROWS = [5,5,5,5,35,35,35,35]
        IFORMROWS = [19,19,19,19,49,49,49]
        FORMCOLS = [7,16,25,34,7,16,25]
        DAYROWS = [3,3,3,3,33,33,33]
        COLWIDTHS = [48,28,28,28,28,30,25,25,25,48,28,28,28,28,30,25,25,25,48,28,28,28,28,30,25,25,25,48,28,28,28,28,30,25,25,25,48,28,28,28,28,30,25,25,25,48,28,28,28,28,30,25,25,25,48,28,28,28,28,30,25,25,25,48,28,28,28,28,30,25,25,25]
        DAYS = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
        DFORMS = ['=ROUNDUP(LARGE(INDIRECT("RC[-5]",0):INDIRECT("RC[-2]",0),1)/$A$1,0)','=ROUNDUP(SMALL(INDIRECT("RC[-6]",0):INDIRECT("RC[-3]",0),1)/$A$1,0)','=ROUNDUP(AVERAGE(INDIRECT("RC[-7]",0):INDIRECT("RC[-4]",0))/$A$1,0)']
        IFORMS = ['=ROUNDUP(LARGE(INDIRECT("RC[-5]",0):INDIRECT("RC[-2]",0),1)/$B$1,0)','=ROUNDUP(SMALL(INDIRECT("RC[-6]",0):INDIRECT("RC[-3]",0),1)/$B$1,0)','=ROUNDUP(AVERAGE(INDIRECT("RC[-7]",0):INDIRECT("RC[-4]",0))/$B$1,0)']
        COLORROWS = [7,21,37,51]
        GREENFILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        BLUEFILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ORANGEFILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        LEFTBORDER = Border(left=Side(border_style='medium',color='FF000000'))
        TOPBORDER = Border(top=Side(border_style='medium',color='FF000000'))
        TOPLEFTBORDER = Border(top=Side(border_style='medium',color='FF000000'),
                                left=Side(border_style='medium',color='FF000000'))
        LEFTCELLS = ['b2','t2','ac2','b3','k3','t3','ac3','b33','k33','t33','b34','k34','t34','b4','k4','t4','ac4']    
        HISTORYCELLS1 = ['b4','k4','t4','ac4','b34','k34','t34']
        HISTORYCELLS2 = ['c4','l4','u4','ad4','c34','l34','u34']
        HISTORYCELLS3 = ['d4','m4','v4','ae4','d34','m34','v34']
        HISTORYCELLS4 = ['e4','n4','w4','af4','e34','n34','w34']

        #Set-up
        ws['a1'] = 3
        ws['b1'] = 15
        ws.sheet_view.showGridLines = False
        for i in range (0,18): #Paste Hours
            for o, hour in enumerate(HOURS):
                ws.cell(row = HOURSROWS[i] + o, column = HOURSCOLS[i], value = hour)
            
        for i in range(0,7):#Paste formulas for high low and average and day names
            
            ws.merge_cells(start_row=DAYROWS[i],end_row = DAYROWS[i],start_column = FORMCOLS[i]-5, end_column = FORMCOLS[i]+2)
            ws.cell(row = DAYROWS[i], column = FORMCOLS[i]-5, value = DAYS[i])
            ws.cell(row=DFORMROWS[i]-1, column = FORMCOLS[i], value = "H").font = Font(bold=True)
            ws.cell(row=DFORMROWS[i]-1, column = FORMCOLS[i]+1, value = "L").font = Font(bold=True)
            ws.cell(row=DFORMROWS[i]-1, column = FORMCOLS[i]+2, value = "A").font = Font(bold=True)
            ws.cell(row=DFORMROWS[i]+13, column = FORMCOLS[i], value = "H").font = Font(bold=True)
            ws.cell(row=DFORMROWS[i]+13, column = FORMCOLS[i]+1, value = "L").font = Font(bold=True)
            ws.cell(row=DFORMROWS[i]+13, column = FORMCOLS[i]+2, value = "A").font = Font(bold=True)
            #DFORMROWS-1
            #FORMCOLS
            for o in range(0,12):
                for p in range(0,3):
                    ws.cell(row = DFORMROWS[i]+o, column = FORMCOLS[i]+p, value = DFORMS[p]).font = Font(bold=True)
                    ws.cell(row = IFORMROWS[i]+o, column = FORMCOLS[i]+p, value = IFORMS[p]).font = Font(bold=True)


        #Set column widths
        for i, col in enumerate(['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p',
                        'q','r','s','t','u','v','w','x','y','z','aa','ab','ac','ad','ae',
                        'af','ag','ah','ai','aj','ak','al','am','an','ao','ap','aq','ar','st']):
            ws.column_dimensions[col].width = COLWIDTHS[i]/7

        #center all cells
        for row in range(1,65):
            for col in range(1,55):
                ws.cell(row = row, column = col).alignment = Alignment(horizontal='center')


        for i in range(1,38):
            for row in COLORROWS:
                ws.cell(row = row, column = i).fill = BLUEFILL
                ws.cell(row = row+3, column = i).fill = GREENFILL
                ws.cell(row = row+7, column = i).fill = ORANGEFILL

        #set font size
        ws['b2'].font = Font(size = 24)
        ws['t2'].font = Font(size = 24)
        ws['ac2'].font = Font(size = 24)
        ws['ac2'] = "Week _"
        ws['t2'] = "Period _"
        ws['a17'] = "Sales"
        ws['a47'] = "Sales"
        #Merge cells
        ws.merge_cells('b2:r2')
        ws.merge_cells('t2:aa2')
        ws.merge_cells('ac2:aj2')
        ws.row_dimensions[2].height = 40.5
        
        #Borders
        for row in range(2,31):
            for col in [2, 10, 11, 19, 20, 28, 29, 37]:
                ws.cell(row = row, column = col).border = LEFTBORDER
        for row in range(33,61):
            for col in [2,10,11,19,20,28]:
                ws.cell(row = row, column = col).border = LEFTBORDER
        
        for row in [2,3,4,31]:
            for col in [2,3,4,5,6,7,8,9,11,12,13,14,15,16,17,18,20,21,22,23,24,25,26,27,29,30,31,32,33,34,35,36]:
                ws.cell(row = row, column = col).border = TOPBORDER
        
        for row in [33,34,61]:
            for col in [2,3,4,5,6,7,8,9,11,12,13,14,15,16,17,18,20,21,22,23,24,25,26,27]:
                ws.cell(row = row, column = col).border = TOPBORDER

        ws['j2'].border = TOPBORDER
        ws['j3'].border = TOPBORDER
        for cell in LEFTCELLS:
            ws[cell].border = TOPLEFTBORDER
        ws['i1'] = "Histories Used:"
        ws['k1'] = self.ui.w1.text()
        ws['l1'] = self.ui.w2.text()
        ws['m1'] = self.ui.w3.text()
        ws['n1'] = self.ui.w4.text()

        for cel1, cel2, cel3, cel4 in zip(HISTORYCELLS1, HISTORYCELLS2, HISTORYCELLS3, HISTORYCELLS4):
            ws[cel1] = "=k1"
            ws[cel2] = "=l1"
            ws[cel3] = "=m1"
            ws[cel4] = "=n1"

        




        del(cell,col,cel1,cel2,cel3,cel4,i,o,p,row,HOURS,HOURSCOLS,HOURSROWS,DFORMROWS,IFORMROWS,FORMCOLS,DAYROWS,COLWIDTHS,DAYS,DFORMS,IFORMS,COLORROWS,GREENFILL,BLUEFILL,ORANGEFILL,LEFTBORDER,TOPBORDER,TOPLEFTBORDER,LEFTCELLS,HISTORYCELLS1,HISTORYCELLS2,HISTORYCELLS3,HISTORYCELLS4)
        SALESFORMULAS = ['=ROUNDUP(LARGE(INDIRECT("RC[-5]",0):INDIRECT("RC[-2]",0),1),0)','=ROUNDUP(SMALL(INDIRECT("RC[-6]",0):INDIRECT("RC[-3]",0),1),0)','=ROUNDUP(AVERAGE(INDIRECT("RC[-7]",0):INDIRECT("RC[-4]",0)),0)']


        for i, file in enumerate(filelist):
            file = openrtf(file)

            ws['b2'] = int(file[1][83:87])
            for row in range(5,17): #Deliveries
                ws.cell(row = row, column = 2+i, value = int(file[row+4][21:25].strip()))#Mon 
                ws.cell(row = row, column = 11+i, value = int(file[row+4][50:54].strip()))#Tue 
                ws.cell(row = row, column = 20+i, value = int(file[row+4][79:83].strip()))#Wed 
                ws.cell(row = row, column = 29+i, value = int(file[row+4][108:112].strip()))#Thur 
                ws.cell(row = row+30, column = 2+i, value = int(file[row+25][21:25].strip()))#Fri
                ws.cell(row = row+30, column = 11+i, value = int(file[row+25][50:54].strip()))#Sat 
                ws.cell(row = row+30, column = 20+i, value = int(file[row+25][79:83].strip()))#Sun 

            for row in range(19,31): #Products
                ws.cell(row = row, column = 2+i, value = int(file[row-10][21+23:25+23].strip()))#Mon 
                ws.cell(row = row, column = 11+i, value = int(file[row-10][50+23:54+23].strip()))#Tue 
                ws.cell(row = row, column = 20+i, value = int(file[row-10][79+23:83+23].strip()))#Wed 
                ws.cell(row = row, column = 29+i, value = int(file[row-10][108+23:112+23].strip()))#Thur 
                ws.cell(row = row+30, column = 2+i, value = int(file[row+11][21+23:25+23].strip()))#Fri
                ws.cell(row = row+30, column = 11+i, value = int(file[row+11][50+23:54+23].strip()))#Sat 
                ws.cell(row = row+30, column = 20+i, value = int(file[row+11][79+23:83+23].strip()))#Sun 

            #Get Sales
            salesline = findline(file, "Net Sales")
            mon = int(round(float(file[salesline][27:34]),-2)//100)
            tue = int(round(float(file[salesline][44:51]),-2)//100)
            wed = int(round(float(file[salesline][61:68]),-2)//100)
            thu = int(round(float(file[salesline][78:85]),-2)//100)
            fri = int(round(float(file[salesline][95:102]),-2)/100)
            sat = int(round(float(file[salesline][112:119]),-2)//100)
            sun = int(round(float(file[salesline][129:136]),-2)//100)
            ws.cell(row=17, column =  1+i+1, value = mon)#Tue
            ws.cell(row=17, column = 10+i+1, value = tue)#Wed
            ws.cell(row=17, column = 19+i+1, value = wed)#Mon
            ws.cell(row=17, column = 28+i+1, value = thu)#Thur
            ws.cell(row=47, column =  1+i+1, value = fri)#Fri
            ws.cell(row=47, column = 10+i+1, value = sat)#Sat
            ws.cell(row=47, column = 19+i+1, value = sun)#Sun
            for j, form in enumerate(SALESFORMULAS):
                ws.cell(row=17, column =  1+j+6, value = form)#Tue
                ws.cell(row=17, column = 10+j+6, value = form)#Wed
                ws.cell(row=17, column = 19+j+6, value = form)#Mon
                ws.cell(row=17, column = 28+j+6, value = form)#Thur
                ws.cell(row=47, column =  1+j+6, value = form)#Fri
                ws.cell(row=47, column = 10+j+6, value = form)#Sat
                ws.cell(row=47, column = 19+j+6, value = form)#Sun

#      Net Sales:         3680.40          2876.97          3232.95          5107.77          6521.32          5671.18          5258.75'

        wb.save(self.outputfolder+"Schedule History.xlsx")
        os.startfile(self.outputfolder+"Schedule History.xlsx")
        self.ui.outputbox.setText("History Report ran successfully, opening output file now")
    except:
        self.ui.outputbox.append("ENCOUNTERED ERROR")
        self.ui.outputbox.append("Please send the contents of this box to Justin")
        self.ui.outputbox.append(traceback.format_exc())
    


if __name__ == '__main__':
    filelist = []
    for file in os.listdir("c:/zocdownload/"):
        if file.startswith("DSHWKC"):
            filelist.append(file)
    filelist = ["c:/zocdownload/"+file for file in filelist]
    run(filelist)
