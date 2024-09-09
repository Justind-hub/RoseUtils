from openpyxl import Workbook
from os import listdir, remove, startfile
import pandas as pd
from time import perf_counter
import traceback

def read_rtf(file2:str)->list[str]:
    with open(file2, 'r') as file:
        text = file.readlines()
    for i in range(0,5):
        text.pop(0)
    text2 = []
    for line in text:
        if len(line) != 5:
            if "\\page" in line:
                text2.append("")
            else:
                text2.append(line[66:-2])
    del(text)
    return text2

def run(self) -> None:
    try:
        wb = Workbook()
        ws = wb.active
        start = perf_counter()
        self.append_text.emit("Running Last Weeks Variances...")                                               
        zocdownload = self.zocdownloadfolder                                                        
        output = self.outputfolder
        delete = self.check_delete                                                                         
        rcp = self.rcp   
        storelist = ["1740","1743","2172","2174","0223","2272","2457","2549","2603","2953","3498","4778"] if rcp else \
                    ["2208","2306","2325","2478","2612","2618","2687","2921","3015","3130","3479","4405","5293"]
        
        storecols = {'1740': 2, '2172': 4,  '2236': 6,  '2272': 8,  '2549': 10, '2953': 12, 
                    '04778': 14, '1743': 16, '2174': 18, '2457': 20, '2603': 22, '3498': 24} if rcp else \
                    {'2208': 2, '2306': 4,  '2325': 6,  '2478': 8,  '2612': 10, '2618': 12, 
                    '2687': 14, '2687': 16, '2921': 18, '3130': 20, '3479': 22, '4405': 24}
                    
        itemdict = {'4045': 'Busboys', '4052': 'Can Liners', '4113': 'Mops', '4120': 'Scrub pad', 
                    '4121': 'Sponges', '4250': 'Toilet Paper', '6000': '20oz Pepsi', 
                    '6001': '20oz Diet', '6003': '20 Mt Dew', '6005': '20oz Mist', 
                    '6006': '20oz Aquifina', '6007': '20oz Dr P', '6009': '20oz Mug', 
                    '6012': '20oz Orange', '6109': '20oz Lifewater', '6200': '2L Pepsi', 
                    '6201': '2L Diet', '6203': '2L Mt Dew', '6205': '2L Mist', '6207': 
                    '2L Dr P', '6209': '2L Mug', '6212': '2L Orange', '1057': 'Ch', 
                    '1121': 'Cream Ch', '1159': '2Ch', '1257': '3Ch', '1313': 'String Ch', 
                    '1074': 'Thin Cr', '1075': '10"', '1077': '12"', '1080': '14"', 
                    '1082': '16"', '1086': 'GF', '1406': 'Cookie', '1407': 'Brownie', 
                    '1505': 'Pullapart', '1040': 'Pe', '1049': 'Ba', '1063': 'Be', 
                    '1065': 'Sa', '1066': 'It', '1071': 'Anchov', '1095': 'Ck', '1098': 'Wings', 
                    '1099': 'Poppers', '1163': 'Salami', '1167': 'Cb', '1178': 'Steak', '1198': 'Mb', 
                    '1016': 'Gp', '1017': 'On', '1019': 'To', '1031': 'Bo', '1048': 'Pi', 
                    '1051': 'Mu', '1052': 'Spinach', '1170': 'Cloves', '1209': 'Bp', '1210': 'Jp', 
                    '1241': 'Peppercini', '1263': 'Ah', '1314': 'Salad', '1002': 'Ranch', 
                    '1006': 'Sauce', '1102': 'Sp Garlic', '1104': 'Garlic jug', '1105': 'Garlic Cup', 
                    '1114': 'Ranch Cup', '1115': 'BBQ Cup', '1116': 'Buff Cup', '1117': 'Hny Must', 
                    '1119': 'Blue Ch', '1122': 'G Parm', '1135': 'Buff', '1140': 'HCP', '1148': 'BBQ', 
                    '1213': 'Ch Cup', '1218': 'Alfredo', '1306': 'Italian Dr', '1308': 'Ranch Dr', 
                    '1501': 'Cream Ch', '1028': 'Dust', '1191': 'Seasoning', '1222': 'Season Pkt', 
                    '1224': 'CRP', '1225': 'Parm', '2005': '10" box', '2007': '12" box', '2010': '14" box', 
                    '2025': '16" box', '2108': 'Pdia box', '2021': '5.5" box', '2031': 'Blaster', 
                    '2036': 'Quality Seal', '2039': 'Foil', '2043': '8" Box', '2047': 'Chicken Box', 
                    '2065': 'Bstick Tray', '2071': 'Knot Tray', '2305': 'Forks', '2306': 'Bowls', 
                    '2307': 'Sleeves', '3020': '2.0 cups', '3021': '2.0 Cups', '3022': 'Lids', 
                    '3040': '14" Parch', '3041': '6.5" Parch', '3042': '10" Parch', '3044': 'Dbag', 
                    '3064': 'Napkins','1014':'Cool Ranch','1002':'Spicy Garlic','1126':'CR Seasoning',
                    '1120':'Lemon Pepper','1506':'Twix','1507':'Caramel','1509':'Oreo','1519':'Cheesecake'}

        filelist = [zocdownload + file for file in listdir(zocdownload) if "INVTAR" in file]
        for file in filelist:
            textlist = read_rtf(file)
            if delete: remove(file)
            sodas, items, dollars, units = [], [], [], []
            store = textlist[0][83:89].strip()
            self.append_text.emit("Running store "+store+"...")                                                ######################################################
            col = storecols[store]
            try:
                storelist.remove(store.strip())
            except:
                pass
            textlist = [line for line in textlist if len(line) == 136]
            textlist.pop(0)
            textlist.pop(0)
            cheese = 0
            
            for line in textlist:

                try:
                    items.append(itemdict[line[4:8]])
                except:
                    items.append(line[11:35].strip())
                dollars.append(float(line[105:116].strip()))
                units.append(float(line[65:79].strip()))
                if line[11:35].strip() == "20lb PIZZA CHEESE": cheese = float(line[105:116].strip())
                elif "20oz" in line[11:35].strip(): sodas.append(float(line[65:79].strip()))
            
            df = pd.DataFrame()
            df["items"],df['dollars'],df['units'] = items, dollars,units
            over = df[(df.dollars > 15)].sort_values(by='dollars',ascending=False)
            under = df[(df.dollars < -15)].sort_values(by='dollars')

            r = 3
            for row in over.iterrows():
                ws.cell(row = r, column = col,value = row[1][0] + " " + str(row[1][1]))
                r+=1

            r = 3
            for row in under.iterrows():
                ws.cell(row = r, column = col+1,value = row[1][0] + " " + str(row[1][1]))
                r+=1

            ws.cell(row=2,column=col,value="Ch " + str(cheese))
            ws.cell(row=2,column=col+1,value=sum(sodas))

        merge = ['b1:c1','d1:e1','f1:g1','h1:i1','j1:k1','l1:m1',
                'n1:o1','p1:q1','r1:s1','t1:u1','v1:w1','y1:z1']

        for item in storecols.items():
            ws.cell(row=1,column=item[1],value=item[0])

        for range in merge:
            ws.merge_cells(range_string=range)

        ws.cell(row=1,column=1,value="Store")
        ws.cell(row=2,column=1,value="Cheese & 20oz")
        wb.save(output+"Last Weeks Variances.xlsx")   
        startfile(f"{output}Last Weeks Variances.xlsx")
        self.append_text.emit(f"Finished {len(filelist)} stores in {round(perf_counter() - start,2)} seconds")
        self.append_text.emit(f"Missing stores {storelist}")
        self.append_text.emit(f"File saved to {output}Last Weeks Variances.xlsx, opening now...")
    except:
        self.append_text.emit("ENCOUNTERED ERROR")
        self.append_text.emit("Please send the contents of this box to Justin")
        self.append_text.emit(traceback.format_exc())
