from striprtf.striprtf import rtf_to_text
from openpyxl import Workbook

from os import startfile, remove
import traceback




def run(self, file) -> None:
    try:



        def openrtf(file): #Call this to open an rtf file with the filepath in the thing.
                            #Will return a list of strings for each line of the file
            with open(file, 'r') as file:
                text = file.read()
            text = rtf_to_text(text)
            textlist = text.splitlines()
            return textlist #returns a list containing entire RTF file


        ITEMNUMLIST = [1002,1006,1016,1017,1019,1028,1031,1040,1048,1049,1051,1052,1056,1063,1065,
                    1066,1071,1074,1075,1077,1080,1082,1086,1095,1098,1099,1102,1104,1105,1114,
                    1115,1116,1117,1119,1122,1135,1140,1148,1159,1163,1167,1170,1178,1191,
                    1198,1209,1210,1213,1218,1222,1224,1225,1241,1257,1263,1306,1308,1313,1314,
                    1406,1407,1501,1505,1509,1519,2005,2007,2010,2047,2025,2039,2039,2043,2065,2071,2108,
                    2305, 2306, 2307, 3020,3022,3040,3041,3042,3044]


        wb = Workbook()
        ws = wb.create_sheet("Export",index = 0) # type: ignore
        textlist = openrtf(file)
        if self.check_delete:
            remove(file)
        textlist = [row for row in textlist if len(row) == 136]
        textlist.pop(0)
        textlist.pop(0)
        itemnums = []
        items = []
        yields = []

        for row in textlist:
            itemnums.append(int(row[4:8]))
            items.append(row[11:39].strip())
            yields.append(float(row[116:127].strip()))

        for i, num in enumerate(ITEMNUMLIST):
            if num not in itemnums:
                continue
            index = itemnums.index(num)
            ws.cell(row = i+1, column = 1, value = num) # type: ignore
            ws.cell(row = i+1, column = 2, value = items[index]) # type: ignore
            ws.cell(row = i+1, column = 3, value = yields[index]) # type: ignore
        if self.gm_filename_checked:
            wb.save(self.outputfolder+self.ui.gm_filename.text()+".xlsx")
            self.ui.outputbox.setText("Yields Report ran successfully, saved to "+self.outputfolder+self.ui.gm_filename.text()+".xlsx")
            startfile(self.outputfolder+self.ui.gm_filename.text()+".xlsx")
        else:
            wb.save(self.outputfolder+"yields.xlsx")
            startfile(self.outputfolder+"yields.xlsx")
            self.ui.outputbox.setText("Yields report Successful, opening output file now")
    except:
        self.ui.outputbox.append("ENCOUNTERED ERROR")
        self.ui.outputbox.append("Please send the contents of this box to Justin")
        self.ui.outputbox.append(traceback.format_exc())
